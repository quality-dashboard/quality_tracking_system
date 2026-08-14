"""
export_to_template.py - 模板导出模块
将系统检测数据填充到用户上传的Excel模板中
保留模板原有格式、公式、合并单元格及列宽
"""
import copy
from io import BytesIO
from typing import Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from data_utils import load_all_records


def _find_data_start_row(ws, header_col: int = 1) -> Optional[int]:
    """
    自动定位模板中数据区域的起始行
    扫描第一列（或指定列），找到第一个非空单元格作为表头行，
    数据从表头行的下一行开始。
    返回数据起始行号（1-based），未找到返回 None。
    """
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=header_col).value
        if cell_value is not None and str(cell_value).strip() != "":
            # 找到表头行，数据从下一行开始
            return row + 1
    return None


def _copy_cell_style(src_cell, dst_cell):
    """
    复制单元格的样式（字体、填充、边框、对齐、数字格式）
    不复制值，仅复制格式
    """
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def fill_template(template_bytes: bytes) -> Tuple[bytes, str]:
    """
    核心导出函数

    Parameters
    ----------
    template_bytes : bytes
        用户上传的 Excel 模板文件的二进制内容

    Returns
    -------
    Tuple[bytes, str]
        (填充后的Excel二进制内容, 状态消息)
    """
    # 1. 读取数据库记录
    df = load_all_records()
    if df.empty:
        return b"", "❌ 数据库中没有检测记录，无法导出。"

    # 2. 加载模板工作簿（保留所有格式）
    try:
        wb = load_workbook(filename=BytesIO(template_bytes))
    except Exception as e:
        return b"", "❌ 模板文件读取失败: {}".format(str(e))

    ws = wb.active

    # 3. 自动定位数据起始行
    data_start_row = _find_data_start_row(ws)
    if data_start_row is None:
        return b"", "❌ 未在模板中找到表头行，请确认模板第一列包含表头信息。"

    # 4. 获取模板列映射
    # 假设模板表头行的列名与 DataFrame 列名一致
    header_row = data_start_row - 1
    col_map = {}  # {df列名: excel列号(1-based)}
    for col_idx in range(1, ws.max_column + 1):
        header_val = ws.cell(row=header_row, column=col_idx).value
        if header_val is not None:
            header_str = str(header_val).strip()
            if header_str in df.columns:
                col_map[header_str] = col_idx

    if not col_map:
        return b"", "❌ 模板表头与数据库字段无匹配项，请检查模板表头名称。"

    # 5. 记录模板行样式（用于扩展时复制格式）
    style_source_row = data_start_row

    # 6. 逐行填充数据
    records = df.to_dict("records")
    total_rows = len(records)

    for i, record in enumerate(records):
        target_row = data_start_row + i

        # 如果超出模板现有行数，插入新行并复制格式
        if target_row > ws.max_row:
            ws.insert_rows(target_row)
            # 从样式源行复制格式到新行
            for col_idx in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=style_source_row, column=col_idx)
                dst_cell = ws.cell(row=target_row, column=col_idx)
                _copy_cell_style(src_cell, dst_cell)

        # 按列映射填入数据
        for col_name, col_idx in col_map.items():
            value = record.get(col_name, "")
            ws.cell(row=target_row, column=col_idx).value = value

    # 7. 保存并返回
    output = BytesIO()
    try:
        wb.save(output)
    except Exception as e:
        return b"", "❌ 生成Excel失败: {}".format(str(e))

    status = "✅ 成功导出 {} 条记录到模板（数据起始行: {}，匹配列: {}个）".format(
        total_rows, data_start_row, len(col_map)
    )
    return output.getvalue(), status